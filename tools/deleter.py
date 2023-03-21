'''

This script deletes all sheets from an Excel file except for the ones specified
Universidad Del Valle de Guatemala


'''


import os
import openpyxl
import xlrd
import xlwt


# Set the directory where the Excel files are stored
directory = '../'

# Loop through each file in the directory
for filename in os.listdir(directory):
    if filename.endswith('.xlsx'):
        # Load the Excel file using openpyxl
        workbook = openpyxl.load_workbook(os.path.join(directory, filename))

        # Loop through each sheet in the workbook
        for sheetname in workbook.sheetnames:
            # Delete any sheets that don't match the specified names
            if sheetname != "Neonatal sexo y causa de muerte" and sheetname != "Edad sexo y causas de muerte":
                workbook.remove(workbook[sheetname])

        # Save the modified workbook
        workbook.save(os.path.join(directory, filename))

    elif filename.endswith('.xls'):
        # Load the Excel file using xlrd
        workbook = xlrd.open_workbook(os.path.join(
            directory, filename), formatting_info=True)

        # Create a new Excel file using xlwt
        new_workbook = xlwt.Workbook()

        # Loop through each sheet in the workbook
        for sheet in workbook.sheets():
            # Check if the sheet name matches the specified names
            if sheet.name == "Neonatal sexo y causa de muerte" or sheet.name == "Edad sexo y causas de muerte":
                # Copy the sheet to the new workbook
                new_sheet = new_workbook.add_sheet(sheet.name)
                for row in range(sheet.nrows):
                    for col in range(sheet.ncols):
                        new_sheet.write(row, col, sheet.cell(row, col).value)

        # Save the modified workbook
        new_workbook.save(os.path.join(directory, filename))
